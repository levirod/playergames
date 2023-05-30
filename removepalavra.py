import openpyxl

def remover_palavras(planilha, palavras_remover):
    contador_linha = 1
    for linha in planilha.iter_rows(min_row=1, max_row=planilha.max_row, values_only=True):
        for indice_coluna, valor in enumerate(linha, start=1):
            if isinstance(valor, str):
                for palavra in palavras_remover:
                    valor = valor.replace(palavra, '')
                planilha.cell(row=contador_linha, column=indice_coluna, value=valor)
        contador_linha += 1

# Abrir o arquivo do Excel
arquivo_excel = r'C:\Users\venda\OneDrive\Área de Trabalho\Resolver Planilha\planilha_simplificada.xlsx'
workbook = openpyxl.load_workbook(arquivo_excel)

# Selecionar a planilha
nome_planilha = 'Worksheet'

try:
    planilha = workbook[nome_planilha]
except KeyError:
    print(f"A planilha '{nome_planilha}' não foi encontrada no arquivo do Excel.")
    workbook.close()
    exit()

# Especificar as palavras a serem removidas
palavras_remover = ['ADRIANOPOLIS - ', 'ASSISTÊNCIA CONSOLE - ', 'CANAL DIGITAL - ', 'CENTRO - ', 'CHAPADA - ', 'DOM PEDRO - ', 'PARQUE DEZ - ', 'VIEIRALVES - '] 

# Chamar a função para remover as palavras específicas
remover_palavras(planilha, palavras_remover)

# Salvar as modificações no arquivo
arquivo_modificado = r'C:\Users\venda\OneDrive\Área de Trabalho\Resolver Planilha\planilha_semlojas.xlsx'
workbook.save(arquivo_modificado)

# Fechar o arquivo do Excel
workbook.close()
